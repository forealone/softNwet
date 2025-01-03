# -*- coding: utf-8 -*-
"""
Created on Thu Sep 23 14:54:04 2021

@author: User
"""

#读一个码表，看看对不对，不对的话可以修改保存

from openpyxl import Workbook
from openpyxl import load_workbook
import re

wb = load_workbook(r'E:\23-个人\月报参照项.xlsx')
ws1 = wb['公司领导排序']
ws2 = wb['干部类别']
ws3 = wb['干部职级']
ws4 = wb['分公司']
ws5 = wb['子公司董监高']

'''
wb = Workbook()
ws1 = wb.active
ws1.append(['储晓明','杨玉成','黄昊','方荣义','卫功琦','张剑','徐宜阳','杨文清','任全胜','刘跃','朱敏杰','薛军','李雪峰','房庆利','王苏龙','吴萌','汤俊','徐亮','何沙','车作斌'])
ws2 = wb.active
ws2.append(['公司正职','公司副职','公司助理','公司总监','总部正职','总部副职','总部助理','二级总部正职','二级总部副职','二级总部助理','分公司正职','分公司副职','分公司助理','子公司正职','子公司副职','子公司助理','总部二级部门经理','分公司二级部门经理','营业部正职','营业部正职（卫星）','营业部副职','营业部助理','非行政职务'])
ws3 = wb.active
ws3.append(['公司正职','公司总经理级','公司副职','公司副总经理级','公司总经理助理级','公司总监级','总部总经理级','总部总经理级待遇','总部副总经理级','总部副总经理级待遇','总部总经理助理级','总部总经理助理级待遇','二级部门经理级','二级部门副经理级','二级部门经理助理级','营业部总经理级','营业部副总经理级','营业部总经理助理级'])
ws4 = wb.active
ws4.append(['上海分公司','江苏分公司','浙江分公司','杭州分公司','北京分公司','四川分公司','深圳分公司','湖北分公司','广东分公司','温州分公司','辽宁分公司','江西分公司','重庆分公司','厦门分公司','大连分公司','宁波分公司','广西分公司','湖南分公司','天津分公司','山东分公司','吉林分公司','安徽分公司','海南分公司','福建分公司','河南分公司','黑龙江分公司','贵州分公司','青岛分公司','陕西分公司','甘肃分公司','山西分公司','宁夏分公司','云南分公司','河北分公司','内蒙古分公司','上海自贸区分公司','沈阳分公司','成都分公司','兰州分公司','西部证券','上海第二分公司','苏州分公司','台州分公司','南通分公司'])
ws5 = wb.active
ws5.append(['李建中','尚恒','毛宗平','马龙官','陈晓升','刘郎','刘震','梁钧','郭纯'])
'''

#输出成一个list，后续编辑
list_ws1 = []
for c in range(1, ws1.max_column+1):
    list_ws1.append(ws1.cell(1,c).value)
print("请校验列表：1、公司领导排序")
print(list_ws1)

#功能1：清空了重来
def fun_01():
    name = input('按序输入新的人员姓名，以空格隔开：')
    while len(name) < 1:
        name = input('不能为空，请重新输入：')
 
    list1 = re.split(r'[\s\,\;\、\，\.\/]+', name)
    print(list1 ,'\n 共',len(list1),'人')
    return list1

#功能2：替换某一个人名
def fun_02(list_origin):
    num = input('输入数字，要修改名单中第几个人：')
    while re.match(r'\d{1,4}', num) == None:
        num = input('输入数字有误，请重新输入：')
    while int(num) > len(list_origin):
        num = input('输入数字超过已有参照项的最大个数，请重新输入：')
    
    list_origin[int(num)-1] = input('按序输入一个人员姓名：')
    print('名单更新如下：')
    print(list_origin)
    return list_origin

#功能3：元素插到指定位置
def fun_03(list_origin):
    name = input('输入一个人员姓名：')
    num = input('插到第几个：')
    while re.match(r'\d{1,4}', num) == None:
        num = input('输入数字有误，请重新输入：')
    while int(int(num)-1) > len(list_origin):
        num = input('输入数字超过已有参照项的最大个数，请重新输入：')

    list_origin.insert(int(num)-1, name)
    print('名单更新如下：')
    print(list_origin)
    return list_origin

#功能4：删除其中一个
def fun_04(list_origin):
    num = input('删除第几个：')
    while re.match(r'\d{1,4}', num) == None:
        num = input('输入数字有误，请重新输入：')
    while int(num) > len(list_origin):
        num = input('输入数字超过已有参照项的最大个数，请重新输入：')

    list_origin.pop(int(num)-1)
    print('名单更新如下：')
    print(list_origin)
    return list_origin


#更新list后重新写到worksheet中
def list_update(ws, list_new):
    ws.delete_rows(1)
    for n in range(1,len(list_new)+1):
        ws.cell(1,n).value = list_new[n-1]

#功能选择
print('功能“1”————清空了重来 \n 功能“2”————替换名单中人名 \n 功能“3”————插入一个人名 \n 功能“4”————删除其中一个人名')
func_choice = input('输入数字，转到对应功能：')
if func_choice == '1':
    list_new = fun_01()
elif func_choice == '2':
    list_new = fun_02(list_ws1)
elif func_choice == '3':
    list_new = fun_03(list_ws1)
elif func_choice == '4':
    list_new = fun_04(list_ws1)
else:
    print("----跳过----")

list_update(ws1, list_new)

wb.save(r'C:\Users\User\Desktop\月报参照项.xlsx')
