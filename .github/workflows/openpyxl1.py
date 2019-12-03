# -*- coding: utf-8 -*-
"""
Created on Tue Oct 15 16:56:21 2019

@author: User
"""


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook(r'E:\1-统计\201910\raw\干部信息明细表（编辑）.xlsx')

wb.guess_types = True  #猜测格式类型
ws = wb.active  #获取第一个sheet
print(ws['A1'].value)



'''
range = 'A1:Y'
range += str(ws.max_row)
tab = Table(displayName='Table1', ref=range)

#第一列套用表头格式、最后一列套用表头格式、隔行变色、隔列变色
style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
tab.tableStyleInfo = style
ws.add_table(tab)
wb.save(r'E:\1-统计\201910\raw\干部信息明细表（编辑）openpyxl.xlsx')
'''

#测试
wb1 = Workbook()    #创建文件对象

ws1 = wb1.active     #获取第一个sheet

ws1['A1'] = 42      #写入数字
ws1['B1'] = "你好"+"automation test" #写入中文（unicode中文也可）

ws1.append([1, 2, 3])    #写入多个单元格

import datetime
import time
ws1['A2'] = datetime.datetime.now()    #写入一个当前时间
#写入一个自定义的时间格式
ws1['A3'] =time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())

wb1.save(r'C:\Users\User\Desktop\test1_openpyxl.xlsx')


