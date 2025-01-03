# -*- coding: utf-8 -*-
"""
Created on Fri Aug  2 16:56:38 2019

@author: User
"""

import pandas as pd
from datetime import datetime
import os
import re
from openpyxl import load_workbook
from countdown import countdown

print('即将开始制作月报-集团和证券公司干部信息明细表...')
seconds_1 = 1
seconds_3 = 3
countdown(seconds_3)

date = '999999'
try:
    with open(r"E:\23-个人\month.txt",'r') as file_to_read:
        s = file_to_read.read()
    exec(s)
except:
    print('未知异常')

while re.match(r'\d{4}(1[0-2]{1}$|0[0-9]{1}$)', date) == None:
    date = input('输入的年月有误，请按格式重新输入6位年月，(格式：YYYYMM):')

print('请检查文件目录是否正确、确保目录下有以下文件：\n “E:\\1-统计\\%s\\raw\\干部信息明细-含分子公司.xlsx \n E:\\1-统计\\%s\\raw\\干部信息明细表（集团模板201911）.xlsx” \n ' %(date,date))
countdown(seconds_3)
if os.access(r'E:\1-统计\%s\raw\干部信息明细-含分子公司.xlsx' %date, os.F_OK):
    pass
else:
    input('【E:\\1-统计\\%s\\raw\\干部信息明细-含分子公司】不存在，是否继续？（按回车键继续...） \n' %date)

if os.access(r'E:\1-统计\%s\raw\干部信息明细表（集团模板201911）.xlsx' %date, os.F_OK):
    pass
else:
    input('【E:\\1-统计\\%s\\raw\\干部信息明细表（集团模板201911）】不存在，是否继续？（按回车键继续...） \n' %date)


cadre_data = pd.read_excel(r'E:\1-统计\%s\raw\干部信息明细-含分子公司.xlsx' %date, sheet_name='sheet1')

'''
cadre_data.head(5)
cadre_data.shape
cadre_data.info
cadre_data.describe
cadre_data.index
cadre_data.values
cadre_data.columns
cadre_data['干部类型']
cadre_data.iloc[1026]
cadre_data.iloc[0,:]
cadre_data.iloc[0,0]
'''

cadre_data.rename(columns={'人员工号':'工号'},inplace=True)
cadre_data.rename(columns={'姓名':'人员姓名'},inplace=True)
cadre_data.rename(columns={'部门':'部门名称'},inplace=True)
cadre_data.rename(columns={'管理人员职级':'干部职级'}, inplace=True)
del cadre_data['组织编码']
del cadre_data['部门编码']
del cadre_data['人员类别']
del cadre_data['一级干部类别']
del cadre_data['干部类型编码']
del cadre_data['职称']
del cadre_data['内部电子邮件']
del cadre_data['手机']

#去重计数和普通计数
num_cadre = len(cadre_data['工号'].drop_duplicates())
count_cadre = cadre_data['工号'].count()
if num_cadre == count_cadre:
    print('已检查：证券公司干部人数 %d 人（一人一条记录） \n' %(num_cadre))
else:
    print('去重人数: %d 人, 不去重人数：%d 人，表格中同一人有多条记录，请检查。\n' %(num_cadre,count_cadre))


#个别职务以[职务]为准
cadre_data['职务'].fillna('-',inplace=True)
print('以下人员[职务]替换[管理人员信息：职务]（非行政职务体现专业通道职务等情形）： \n', cadre_data.loc[cadre_data[cadre_data['职务'].str.contains('董事副总经理|业务董事|总部副总经理级待遇')].index,['部门名称','人员姓名','职务','管理人员信息：职务']])
countdown(seconds_1)
job_y = cadre_data.loc[cadre_data[cadre_data['职务'].str.contains('董事副总经理|业务董事|挂职甘肃|总部副总经理级待遇')].index,['职务']]
cadre_data.loc[cadre_data[cadre_data['职务'].str.contains('董事副总经理|业务董事|总部副总经理级待遇')].index,['管理人员信息：职务']] = job_y.iloc[:,0]

#其他干部职务以[管理人员信息：职务]为准
cadre_data['职务比对'] = 'NaN'
cadre_data.loc[cadre_data[cadre_data['职务']==cadre_data['管理人员信息：职务']].index,['职务比对']] = '相同'
print('以下人员职务信息将以[管理人员信息：职务]为准： \n', cadre_data.loc[cadre_data[cadre_data['职务比对']=='NaN'].index,['人员姓名','部门名称','职务','管理人员信息：职务']])
countdown(seconds_1)

del cadre_data['职务']
del cadre_data['职务比对']
cadre_data.rename(columns={'管理人员信息：职务':'职务'},inplace=True)

#因个别干部人事关系在证券公司，但实际任职职务在子公司，需修改对应人员的组织、部门字段
print('因个别干部人事关系在证券公司（其他部门-外派），但实际任职职务在子公司，需修改对应人员的组织、部门字段 \n')

print('夏明睿体现申万宏源（香港）有限公司总经理助理 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='夏明睿'].index,['组织','一级部门','部门名称']] = ['申万宏源（香港）有限公司','申万宏源（香港）有限公司','申万宏源（香港）有限公司']

print('戴佳明体现申银万国投资有限公司总经理 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='戴佳明'].index,['组织','一级部门','部门名称']] = ['申银万国投资有限公司','申银万国投资有限公司','申银万国投资有限公司']

print('邓伟体现服务公司总经理助理 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='邓伟'].index,['组织','一级部门','部门名称']] = ['非一体化管控子公司','上海申银万国综合服务有限公司','上海申银万国综合服务有限公司']

print('黄伟主职部门为甘肃分公司 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['工号']=='360001'].index,['一级部门','部门名称']] = ['甘肃分公司','甘肃分公司']

print('吴华丕体现申万宏源新加坡私人有限公司董事、行政总裁（CEO） \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='吴华丕'].index,['组织','一级部门','部门名称']] = ['申万宏源新加坡私人有限公司','申万宏源新加坡私人有限公司','申万宏源新加坡私人有限公司']

print('谢晨管理人员属性体现为空值（临时）会导致统计表中少一人 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='谢晨'].index,['干部类型']] = ['']

print('俞力黎相应的部门体现申万宏源（香港）有限公司 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='俞力黎'].index,['组织','一级部门','部门名称']] = ['申万宏源（香港）有限公司','申万宏源（香港）有限公司','申万宏源（香港）有限公司']

print('梁钧相应的部门体现申万宏源（香港）有限公司 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='梁钧'].index,['组织','一级部门','部门名称']] = ['申万宏源（香港）有限公司','申万宏源（香港）有限公司','申万宏源（香港）有限公司']

print('黄荣相应的部门体现申万宏源（香港）有限公司 \n')
countdown(seconds_1)
cadre_data.loc[cadre_data[cadre_data['人员姓名']=='黄荣'].index,['组织','一级部门','部门名称']] = ['申万宏源（香港）有限公司','申万宏源（香港）有限公司','申万宏源（香港）有限公司']

'''贺添（年龄每年得手动修改一下）
print('香港公司贺添是经过公司党委面试和任命的干部，但是信息在人力系统中没有，为其手动添加一条记录 \n')
input('按回车确认修改... \n')
insertrow1 = pd.DataFrame({'组织':['非一体化管控子公司'],'一级部门':['申万宏源（香港）有限公司'],'部门名称':['申万宏源（香港）有限公司'],'工号':['--'],'人员姓名':['贺添'],'性别':['男'],'职务':['副总经理'],'干部类型':['子公司副职'],'干部职级':['总部副总经理级'],'政治面貌':['群众'],'最高学历':['硕士研究生毕业'],'出生日期':['1980-05'],'年龄':['41'],'身份证号':[''],'职等':[''],'本岗位任职日期':['2018-12-14'],'现职级任职日期':['2018-12-14']})

cadre_data = cadre_data.append(insertrow1,ignore_index=True,sort=False)
'''

#集团公司干部拼表(注意修改日期)
print('\n 即将把集团公司干部信息拼入表格，请确保以下目录正确、集团公司干部信息表存在')
print('E:\\1-统计\\%s\\raw\\干部信息明细表（集团模板201911）.xlsx' %date)
countdown(seconds_3)
group_cadre_data = pd.read_excel(r'E:\1-统计\%s\raw\干部信息明细表（集团模板201911）.xlsx' %date, skiprows=1, sheet_name='Sheet1')
'''
#导入excel时不用skiprows=1的话，需要做如下处理
p_data.rename(columns={'集团公司干部信息明细表':'部门名称','Unnamed: 1':'人员工号','Unnamed: 2':'人员姓名','Unnamed: 3':'性别','Unnamed: 4':'职务','Unnamed: 5':'干部类型','Unnamed: 6':'干部职级','Unnamed: 7':'政治面貌','Unnamed: 8':'最高学历','Unnamed: 9':'职称','Unnamed: 10':'出生日期','Unnamed: 11':'年龄','Unnamed: 12':'身份证号','Unnamed: 13':'职等','Unnamed: 14':'当前岗位任职开始时间','Unnamed: 15':'手机'},inplace=True)
p_data.drop([0,0],inplace=True)
p_data.reset_index(drop=True, inplace=True)
'''

#计数
num_group_cadre = len(group_cadre_data['人员工号'].drop_duplicates())
count_group_cadre = group_cadre_data['人员工号'].count()
if num_group_cadre == count_group_cadre:
    print('已检查：集团公司干部人数 %d 人（一人一条记录） \n' %(num_group_cadre))
else:
    print('去重人数: %d 人, 不去重人数：%d 人，表格中同一人有多条记录，请检查。\n' %(num_group_cadre,count_group_cadre))

group_cadre_data.insert(0, '组织',None)
group_cadre_data.insert(1, '一级部门',group_cadre_data['部门名称'])
group_cadre_data.loc[group_cadre_data[group_cadre_data['部门名称']=='宏源汇富创业投资有限公司'].index,['组织','一级部门']] = ['宏源汇富创业投资有限公司','宏源汇富创业投资有限公司']
group_cadre_data.loc[group_cadre_data[group_cadre_data['部门名称']=='宏源汇智投资有限公司'].index,['组织','一级部门']] = ['宏源汇智投资有限公司','宏源汇智投资有限公司']
group_cadre_data.loc[group_cadre_data[group_cadre_data['部门名称']=='宏源期货有限公司'].index,['组织','一级部门']] = ['宏源期货有限公司','宏源期货有限公司']
group_cadre_data['组织'].fillna('申万宏源集团股份有限公司',inplace=True)

group_cadre_data.rename(columns={'人员工号':'工号'},inplace=True)
del group_cadre_data['职称']
del group_cadre_data['手机']

cadre_data = pd.concat([group_cadre_data,cadre_data],axis=0,ignore_index=True,sort=False)

#政治面貌和最高学历分类（pivot table用）
cadre_data['政治面貌分类'] = '群众'
cadre_data.loc[cadre_data[cadre_data['政治面貌'].str.contains('民盟|民革|民建|民进|农工党|致公党|九三学社|民主自治同盟')].index,['政治面貌分类']] = '民主党派'
cadre_data.loc[cadre_data[cadre_data['政治面貌'].str.contains('中共')].index,['政治面貌分类']] = '中共党员'

cadre_data['最高学历分类'] = '硕士研究生'
cadre_data['最高学历'].fillna('硕士研究生',inplace=True)
cadre_data.loc[cadre_data[cadre_data['最高学历'].str.contains('专科|中专|高中|初中|小学|职高|技校|中技')].index,['最高学历分类']] = '大学专科及以下'
cadre_data.loc[cadre_data[cadre_data['最高学历'].str.contains('本科')].index,['最高学历分类']] = '大学本科'
cadre_data.loc[cadre_data[cadre_data['最高学历'].str.contains('硕士|研究生毕业班')].index,['最高学历分类']] = '硕士研究生'
cadre_data.loc[cadre_data[cadre_data['最高学历'].str.contains('博士')].index,['最高学历分类']] = '博士研究生'

#年龄分段
cadre_data['年龄段']='-'
cadre_data['年龄'] = cadre_data['年龄'].astype('int')
cadre_data.loc[cadre_data[(cadre_data['年龄']<=35)].index,['年龄段']] = '35岁及以下'
cadre_data.loc[cadre_data[(cadre_data['年龄']>35)&(cadre_data['年龄']<=45)].index,['年龄段']] = '36-45岁'
cadre_data.loc[cadre_data[(cadre_data['年龄']>45)].index,['年龄段']] = '45岁以上'

#取出生年份与当前年度比较，筛选出拟退岗/退休的
cadre_data['出生年份'] = None
cadre_data['退岗标识'] = None
cadre_data['退休标识'] = None
dt = datetime.now()
cadre_data['基准年份'] = int(dt.strftime('%Y')) #转int型以便计算

#出生日期取年份，个别字段出生日期为YYYY-MM的，先统一到YYYY-MM-DD
i = 0
while i < cadre_data.shape[0]:
    if len(cadre_data.loc[i, '出生日期']) == 7:
        cadre_data.loc[i, '出生日期'] = cadre_data.loc[i, '出生日期'] + '-01'
    else:
        pass
    cadre_data.loc[i, '出生年份'] = datetime.strftime(datetime.strptime(cadre_data.loc[i, '出生日期'], '%Y-%m-%d'), '%Y')
    i = i + 1

cadre_data['出生年份']=cadre_data['出生年份'].astype('int')

#拟退休人员打标记
cadre_data.loc[cadre_data[(cadre_data['性别'] == '男') & (cadre_data['基准年份'] - cadre_data['出生年份'] >= 59) | 
        (cadre_data['性别'] == '女') & (cadre_data['基准年份'] - cadre_data['出生年份'] >= 54)].index, '退休标识'] = '1年内退休'

#拟退岗人员打标记（1年内退休的人不再体现退岗标识）
cadre_data.loc[cadre_data[(cadre_data['性别'] == '男') & (cadre_data['基准年份'] - cadre_data['出生年份'] >= 57) & (cadre_data['基准年份'] - cadre_data['出生年份'] < 59) | 
        (cadre_data['性别'] == '女') & (cadre_data['基准年份'] - cadre_data['出生年份'] >= 52) & (cadre_data['基准年份'] - cadre_data['出生年份'] < 54)].index, '退岗标识'] = '到退岗年龄'
#预览一下有哪些拟退岗、退休的
cadre_data.loc[cadre_data[cadre_data['退休标识'] == '1年内退休'].index, ['人员姓名','部门名称','职务']]
cadre_data.loc[cadre_data[(cadre_data['退岗标识'] == '到退岗年龄')].index, ['人员姓名','部门名称','职务']]

#拼考核结果(注意每年更新考核结果，修改日期——将来更新自动抓取近两年文件)
print('确保目录下有以下文件：\n “E:\\2-年度考核\\历年考核结果\\2022考核结果汇总.xlsx  \n E:\\2-年度考核\\历年考核结果\\2023考核结果汇总.xlsx”... \n')
countdown(seconds_1)
if os.access(r'E:\2-年度考核\历年考核结果\2022考核结果汇总.xlsx', os.F_OK):
    print('即将匹配2022年度考核结果... \n')
    countdown(seconds_1)
else:
    input('【E:\\2-年度考核\\历年考核结果\\2022考核结果汇总.xlsx】不存在，是否继续？（按回车键继续...） \n')
if os.access(r'E:\2-年度考核\历年考核结果\2023考核结果汇总.xlsx', os.F_OK):
    print('即将匹配2023年度考核结果... \n')
    countdown(seconds_1)
else:
    input('【E:\\2-年度考核\\历年考核结果\\2023考核结果汇总.xlsx】不存在，是否继续？（按回车键继续...） \n')
    
eva_2022 = pd.read_excel(r'E:\2-年度考核\历年考核结果\2022考核结果汇总.xlsx',sheet_name='干部员工')
eva_2023 = pd.read_excel(r'E:\2-年度考核\历年考核结果\2023考核结果汇总.xlsx',sheet_name='干部员工')
del eva_2022['职务']
del eva_2023['职务']
del eva_2022['序号']
del eva_2023['序号']
del eva_2022['考核对象']
del eva_2023['考核对象']
del eva_2022['部门']
del eva_2023['部门']
del eva_2022['考核方案']
del eva_2023['考核方案']
del eva_2022['导入/同步']
del eva_2023['导入/同步']
del eva_2022['备注']
del eva_2023['备注']
eva_2022.rename(columns={'考核对象编码':'工号'},inplace=True)
eva_2023.rename(columns={'考核对象编码':'工号'},inplace=True)
eva_2022.rename(columns={'考核结果':'2022年度考核结果'},inplace=True)
eva_2023.rename(columns={'考核结果':'2023年度考核结果'},inplace=True)

cadre_data = pd.merge(cadre_data,eva_2022,on='工号',how='left')
cadre_data = pd.merge(cadre_data,eva_2023,on='工号',how='left')
cadre_data['2022年度考核结果'].fillna('-',inplace=True)
cadre_data['2023年度考核结果'].fillna('-',inplace=True)

cadre_data['考核结果是否符合提聘条件'] = None
cadre_data.loc[cadre_data[(cadre_data['2022年度考核结果'].str.contains('A|B')) & (cadre_data['2023年度考核结果'].str.contains('A|B|C'))].index,['考核结果是否符合提聘条件']] = '符合'
cadre_data.loc[cadre_data[(cadre_data['2022年度考核结果'].str.contains('A|B|C')) & (cadre_data['2023年度考核结果'].str.contains('A|B'))].index,['考核结果是否符合提聘条件']] = '符合'
cadre_data['考核结果是否符合提聘条件'].fillna('不符合',inplace=True)

#有人一年有多个结果，导致出现重复记录，需处理一下（只保留第一条记录）
cadre_data.drop_duplicates(['工号'], keep='first', inplace=True)
cadre_data.reset_index(drop=True, inplace=True)

#总部级干部类别（集团公司总部，证券公司事业部总部，分公司本部，子公司、营业部）
cadre_data['总部级干部类别'] = None
cadre_data.loc[cadre_data[(cadre_data['组织'] == '申万宏源集团股份有限公司') & (cadre_data['干部类型'].str.contains('总部正职|总部副职|总部助理'))].index,['总部级干部类别']] = '集团公司总部领导班子'
cadre_data.loc[cadre_data[(cadre_data['组织'] == '申万宏源集团股份有限公司') & (cadre_data['干部类型'].str.contains('总部二级部门经理'))].index,['总部级干部类别']] = '集团公司总部二级部门经理'
cadre_data.loc[cadre_data[((cadre_data['组织'] == '申万宏源证券有限公司') | (cadre_data['组织'].str.contains('承销保荐|上海第二分公司|北京第二分公司|资管公司|资产管理有限公司'))) & (cadre_data['干部类型'].str.contains('总部正职|总部副职|总部助理'))].index,['总部级干部类别']] = '证券公司事业部总部领导班子'
cadre_data.loc[cadre_data[((cadre_data['组织'] == '申万宏源证券有限公司') | (cadre_data['组织'].str.contains('承销保荐|上海第二分公司|北京第二分公司|资管公司|资产管理有限公司'))) & (cadre_data['干部类型'].str.contains('总部二级部门经理'))].index,['总部级干部类别']] = '证券公司事业部总部二级部门经理'
cadre_data.loc[cadre_data[((cadre_data['组织'].str.contains('分公司')) | (cadre_data['组织'] == '申万宏源西部证券有限公司')) & (~cadre_data['组织'].str.contains('上海第二分公司|北京第二分公司')) & (cadre_data['干部类型'].str.contains('分公司正职|分公司副职|分公司助理')) & (~cadre_data['一级部门'].str.contains('营业部'))].index,['总部级干部类别']] = '分公司本部领导班子'
cadre_data.loc[cadre_data[((cadre_data['组织'].str.contains('分公司')) | (cadre_data['组织'] == '申万宏源西部证券有限公司')) & (~cadre_data['组织'].str.contains('上海第二分公司|北京第二分公司')) & (cadre_data['干部类型'].str.contains('分公司二级部门经理')) & (~cadre_data['一级部门'].str.contains('营业部'))].index,['总部级干部类别']] = '分公司二级部门经理'
cadre_data.loc[cadre_data[cadre_data['干部类型'].str.contains('子公司正职|子公司副职|子公司助理')].index,['总部级干部类别']] = '子公司领导班子'


#部门类别（总部业务部门、总部职能部门、总部党群部门、分公司、子公司、营业部）——未独立履职分公司计入营业部，根据实际情况随时调整
cadre_data['部门类别'] = '其他' #先全部设定为“其他”，按照以下的逻辑，新增/更名的部门，以及非行政职务的干部，其部门类别会设定为“其他”
cadre_data.loc[cadre_data[cadre_data['一级部门'].str.contains('零售客户事业部|财富管理事业部|机构客户事业部|机构客户总部|投资交易事业部|国际业务总部|资产管理事业部|固定收益交易总部|固定收益外汇商品事业部（FICC事业部）|固定收益销售交易总部|固定收益融资总部|场外市场总部|金融创新总部|证券投资总部|证券投资事业部|证券金融总部|承销保荐|多元金融部|投资管理部|产业投资管理子公司|托管中心|申万宏源证券资产管理有限公司')].index,['部门类别']] = '总部业务部门'
cadre_data.loc[cadre_data[cadre_data['一级部门'].str.contains('监事会办公室|办公室|计划财务管理总部|董事会办公室|法律合规总部|风险管理总部|内核评审总部|风险资产处置办公室|信息技术保障总部|信息技术开发总部|信息技术架构组|审计总部|运营中心|战略规划总部|战略客户总部|资金营运总部|其他|计划财务部|人力资源部|战略管理部|总经理办公室|稽核审计部|法务风控部|资产运营总部|行政管理总部|上海第二分公司|北京第二分公司')].index,['部门类别']] = '总部职能部门'
cadre_data.loc[cadre_data[cadre_data['一级部门'].str.contains('扶贫办公室|帮扶办公室|纪检监察室|纪律检查室|党委办公室|党委巡视办公室|工会办公室|团委办公室|党委组织部/人力资源总部|党委巡察办公室|党建工作部|企业文化部|纪委')].index,['部门类别']] = '总部党群部门'
cadre_data.loc[cadre_data[(cadre_data['一级部门'].str.contains('分公司')) & (~cadre_data['一级部门'].str.contains('上海第二分公司|北京第二分公司')) | (cadre_data['一级部门'] == '西部证券')].index,['部门类别']] = '分公司'
cadre_data.loc[cadre_data[cadre_data['一级部门'].str.contains('申万菱信|申银万国期货|申银万国证券研究所|申银万国投资|申银万国创新证券投资|申万宏源（香港）|申万宏源（国际）|申万宏源新加坡|宏源汇智|宏源汇富|宏源期货|综合服务有限公司')].index,['部门类别']] = '子公司'
cadre_data.loc[cadre_data[cadre_data['一级部门'].str.contains('营业部')].index,['部门类别']] = '营业部'
cadre_data.loc[cadre_data[cadre_data['一级部门'] == '公司领导'].index,['部门类别']] = '公司领导'
cadre_data.loc[cadre_data[cadre_data['干部类型'] == '非行政职务'].index,['部门类别']] = '其他'
cadre_data.loc[cadre_data[cadre_data['干部类型'] == '转岗人员'].index,['部门类别']] = '其他'
cadre_data.loc[cadre_data[cadre_data['干部类型'] == '转任干部'].index,['部门类别']] = '其他'


#公司领导类别（党委班子）（证券经营班子、集团经营班子）——根据实际情况随时调整
cadre_data['公司领导类别1'] = None
cadre_data.loc[cadre_data[(cadre_data['人员姓名'].isin(['刘健','黄昊','方荣义','卫功琦','张剑'])) & (cadre_data['一级部门']=='公司领导')].index,['公司领导类别1']] = '集团和证券公司党委班子'
cadre_data['公司领导类别2'] = None
cadre_data.loc[cadre_data[(cadre_data['人员姓名'].isin(['黄昊','任全胜','刘跃'])) & (cadre_data['一级部门']=='公司领导')].index,['公司领导类别2']] = '集团公司经营班子'
cadre_data.loc[cadre_data[(cadre_data['人员姓名'].isin(['张剑','李雪峰','陈秀清','王苏龙','吴萌','汤俊','张翼飞','周海晨'])) & (cadre_data['一级部门']=='公司领导')].index,['公司领导类别2']] = '证券公司经营班子'

#拆分重组与排序
if os.access(r'E:\23-个人\月报参照项.xlsx', os.F_OK):
    pass
else:
    input('【E:\\23-个人\\月报参照项.xlsx】不存在，是否继续？（按回车键继续...） \n')

wb = load_workbook(r'E:\23-个人\月报参照项.xlsx')
ws1 = wb['公司领导排序']
ws2 = wb['干部类别']
ws3 = wb['干部职级']
ws4 = wb['分公司']
ws5 = wb['子公司董监高']

def read_list(ws):
    list_sorted = []
    for c in range(1, ws.max_column+1):
        list_sorted.append(ws.cell(1,c).value)
    return list_sorted

list_sorted1 = read_list(ws1)
list_sorted2 = read_list(ws2)
list_sorted3 = read_list(ws3)
list_sorted4 = read_list(ws4)
list_sorted5 = read_list(ws5)

#分拆表1：公司领导
cadre_data_p1 = cadre_data.iloc[cadre_data[cadre_data['部门类别'] == '公司领导'].index,:].reset_index(drop=True)
cadre_data_p1['人员姓名cat'] = cadre_data_p1['人员姓名'].astype('category')
cadre_data_p1['人员姓名cat'].cat.set_categories(list_sorted1, inplace=True)
cadre_data_p1 = cadre_data_p1.sort_values('人员姓名cat', ascending=True)
cadre_data_p1 = cadre_data_p1.reset_index(drop=True)
del cadre_data_p1['人员姓名cat']
#分拆表2：集团和证券公司总部
cadre_data_p2 = cadre_data.iloc[cadre_data[(cadre_data['部门类别'].isin(['总部业务部门','总部职能部门','总部党群部门']))&(cadre_data['干部类型'] !='非行政职务')&(cadre_data['干部类型'] !='转岗人员')&(cadre_data['干部类型'] !='转任干部')].index,:].reset_index(drop=True)

cadre_data_p2['干部类型cat'] = cadre_data_p2['干部类型'].astype('category')
cadre_data_p2['干部类型cat'].cat.set_categories(list_sorted2, inplace=True)

cadre_data_p2['干部职级cat'] = cadre_data_p2['干部职级'].astype('category')
cadre_data_p2['干部职级cat'].cat.set_categories(list_sorted3, inplace=True)

cadre_data_p2 = cadre_data_p2.sort_values(['组织','部门类别','一级部门','干部类型cat','干部职级cat','现职级任职日期','本岗位任职日期'],ascending=(False,True,False,True,True,True,True))
cadre_data_p2 = cadre_data_p2.reset_index(drop=True)
del cadre_data_p2['干部类型cat']
del cadre_data_p2['干部职级cat']
#分拆表3：分公司本部
cadre_data_p3 = cadre_data.iloc[cadre_data[(cadre_data['部门类别']=='分公司')&(cadre_data['干部类型'] !='非行政职务')&(cadre_data['干部类型'] !='转岗人员')&(cadre_data['干部类型'] !='转任干部')].index,:].reset_index(drop=True)
cadre_data_p3['一级部门cat'] = cadre_data_p3['一级部门'].astype('category')
cadre_data_p3['一级部门cat'].cat.set_categories(list_sorted4, inplace=True)

cadre_data_p3['干部类型cat'] = cadre_data_p3['干部类型'].astype('category')
cadre_data_p3['干部类型cat'].cat.set_categories(list_sorted2, inplace=True)

cadre_data_p3['干部职级cat'] = cadre_data_p3['干部职级'].astype('category')
cadre_data_p3['干部职级cat'].cat.set_categories(list_sorted3, inplace=True)

cadre_data_p3 = cadre_data_p3.sort_values(['一级部门cat','干部类型cat','干部职级cat','现职级任职日期','本岗位任职日期'],ascending=(True,True,True,True,True))
cadre_data_p3 = cadre_data_p3.reset_index(drop=True)
del cadre_data_p3['一级部门cat']
del cadre_data_p3['干部类型cat']
del cadre_data_p3['干部职级cat']
#分拆表4：子公司
cadre_data_p4 = cadre_data.iloc[cadre_data[(cadre_data['部门类别']=='子公司')&(cadre_data['干部类型'] !='非行政职务')&(cadre_data['干部类型'] !='转岗人员')&(cadre_data['干部类型'] !='转任干部')].index,:].reset_index(drop=True)
cadre_data_p4['干部类型cat'] = cadre_data_p4['干部类型'].astype('category')
cadre_data_p4['干部类型cat'].cat.set_categories(list_sorted2, inplace=True)

cadre_data_p4['董监高cat'] = cadre_data_p4['人员姓名'].astype('category')
cadre_data_p4['董监高cat'].cat.set_categories(list_sorted5, inplace=True)

cadre_data_p4['干部职级cat'] = cadre_data_p4['干部职级'].astype('category')
cadre_data_p4['干部职级cat'].cat.set_categories(list_sorted3, inplace=True)

cadre_data_p4 = cadre_data_p4.sort_values(['一级部门','干部类型cat','董监高cat','干部职级cat','现职级任职日期','本岗位任职日期'],ascending=(False,True,True,True,True,True))
cadre_data_p4 = cadre_data_p4.reset_index(drop=True)
del cadre_data_p4['干部类型cat']
del cadre_data_p4['董监高cat']
del cadre_data_p4['干部职级cat']
#分拆表5：营业部
cadre_data_p5 = cadre_data.iloc[cadre_data[(cadre_data['部门类别']=='营业部')&(cadre_data['干部类型'] !='非行政职务')&(cadre_data['干部类型'] !='转岗人员')&(cadre_data['干部类型'] !='转任干部')].index,:].reset_index(drop=True)
cadre_data_p5.loc[:,'组织new'] = cadre_data_p5['组织'].str.replace('申万宏源证券有限公司','') 
cadre_data_p5.loc[:,'组织new'] = cadre_data_p5['组织new'].str.replace('申万宏源西部证券有限公司','西部证券') 
cadre_data_p5.loc[:,'组织new'] = cadre_data_p5['组织new'].str.replace('西部证券甘肃分公司','甘肃分公司') 
cadre_data_p5.loc[:,'组织new'] = cadre_data_p5['组织new'].str.replace('西部证券陕西分公司','陕西分公司') 
cadre_data_p5['组织cat'] = cadre_data_p5['组织new'].astype('category')
cadre_data_p5['组织cat'].cat.set_categories(list_sorted4, inplace=True)

cadre_data_p5['干部类型cat'] = cadre_data_p5['干部类型'].astype('category')
cadre_data_p5['干部类型cat'].cat.set_categories(list_sorted2, inplace=True)

cadre_data_p5['干部职级cat'] = cadre_data_p5['干部职级'].astype('category')
cadre_data_p5['干部职级cat'].cat.set_categories(list_sorted3, inplace=True)

cadre_data_p5 = cadre_data_p5.sort_values(['组织cat','一级部门','干部类型cat','干部职级cat','现职级任职日期','本岗位任职日期'],ascending=(True,True,True,True,True,True))
cadre_data_p5 = cadre_data_p5.reset_index(drop=True)
del cadre_data_p5['组织cat']
del cadre_data_p5['组织new']
del cadre_data_p5['干部类型cat']
del cadre_data_p5['干部职级cat']
#分拆表6：非行政职务、转任干部以及其他情况（新设部门没有及时更新脚本中的部门类别的话，相关人员会出现在这里）
cadre_data_p6 = cadre_data.iloc[cadre_data[(cadre_data['部门类别'] =='其他')&(cadre_data['干部类型'] != '转岗人员')].index,:].reset_index(drop=True)

cadre_data_p6 = cadre_data_p6.sort_values(by=['干部类型','组织'],ascending=(True,True))
cadre_data_p6 = cadre_data_p6.reset_index(drop=True)
#分拆表7：转岗人员（不拼进总表）
cadre_data_p7 = cadre_data.iloc[cadre_data[(cadre_data['部门类别'] =='其他')&(cadre_data['干部类型'] == '转岗人员')].index,:].reset_index(drop=True)

cadre_data_p7 = cadre_data_p7.sort_values(by=['干部类型','组织'],ascending=(True,True))
cadre_data_p7 = cadre_data_p7.reset_index(drop=True)

#6张表拼表
cadre_data_sort = cadre_data_p1.append(cadre_data_p2,ignore_index=True)
cadre_data_sort = cadre_data_sort.append(cadre_data_p3,ignore_index=True)
cadre_data_sort = cadre_data_sort.append(cadre_data_p4,ignore_index=True)
cadre_data_sort = cadre_data_sort.append(cadre_data_p5,ignore_index=True)
cadre_data_sort = cadre_data_sort.append(cadre_data_p6,ignore_index=True)
cadre_data = cadre_data_sort

'''
#管理人员属性职务替换职务字段内容（按工号）
def job2to1(x):
    name_x = cadre_data.loc[cadre_data[cadre_data['工号']==x].index,['人员姓名']]
    job_x = cadre_data.loc[cadre_data[cadre_data['工号']==x].index,['职务']]
    job2_x = cadre_data.loc[cadre_data[cadre_data['工号']==x].index,['管理人员信息：职务']]
    print('干部%s（%s）职务将由: \n【%s】 \n更新为: \n【%s】\n' %(name_x.iloc[0,0],x,job_x.iloc[0,0],job2_x.iloc[0,0]))
    cadre_data.loc[cadre_data[cadre_data['工号']==x].index,['职务']] = job2_x.iloc[0,0]
    if job2_x.iloc[0,0] == cadre_data.loc[cadre_data[cadre_data['工号']==x].index,['职务']].iloc[0,0]:
        print('更新完成')
    else:
        print('未更新成功，请检查干部名单中是否有此人')
    print(name_x)
'''
#去重计数和普通计数(再次校验一下，拼表以后有没有人因为排序原因丢失)

num_cadre_total = len(cadre_data['工号'].drop_duplicates())
count_cadre_total = cadre_data['工号'].count()
if num_cadre_total == count_cadre_total:
    print('应有集团和证券各级干部',num_cadre,'+',num_group_cadre,' = ',num_cadre+num_group_cadre,'人')
    print('实有',num_cadre_total,'人，人数符合 \n')
else:
    print('去重人数: %d 人, 不去重人数：%d 人，表格中同一人有多条记录，请检查。\n' %(num_cadre_total,count_cadre_total))


#输出
print('将输出文件至目录E:\\1-统计\\%s\\raw\\' %date)
countdown(seconds_1)

cadre_data_output = pd.ExcelWriter(r'E:\1-统计\%s\raw\干部信息明细表（数据清洗）.xlsx' %date)
cadre_data.to_excel(cadre_data_output, sheet_name='干部信息明细表总表', index=False)
cadre_data_p1.to_excel(cadre_data_output, sheet_name='集团和证券公司领导', index=False)
cadre_data_p2.to_excel(cadre_data_output, sheet_name='集团和证券总部干部', index=False)
cadre_data_p3.to_excel(cadre_data_output, sheet_name='分公司干部', index=False)
cadre_data_p4.to_excel(cadre_data_output, sheet_name='子公司干部', index=False)
cadre_data_p5.to_excel(cadre_data_output, sheet_name='营业部干部', index=False)
cadre_data_p6.to_excel(cadre_data_output, sheet_name='其他干部', index=False)
cadre_data_p7.to_excel(cadre_data_output, sheet_name='转岗人员', index=False)
cadre_data_output.save()

#输出一张给财富的分公司营业部干部表
cadre_data_sort_other = cadre_data_p3.append(cadre_data_p5,ignore_index=True)
cadre_data_output2 = pd.ExcelWriter(r'E:\1-统计\%s\raw\分支机构干部名单（raw）to韩冰.xlsx' %date)
cadre_data_sort_other.to_excel(cadre_data_output2, sheet_name='分公司营业部干部明细', index=False)
cadre_data_output2.save()