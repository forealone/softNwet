# -*- coding: utf-8 -*-
"""
Created on Wed Nov 27 13:47:55 2019

@author: User
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment

wb = load_workbook(r'E:\组织部共享\干部结构统计表\本年度干部变动统计-截止201911.xlsx')

ws1 = wb['干部变动明细']
ws2 = wb['透视图']
ws3 = wb['文字描述']

#字体
font1 = Font(name='宋体', color='000000', size=10, b=True)
font2 = Font(name='宋体', color='000000', size=10)
font3 = Font(name='黑体', color='000000', size=14, b=True)
#单元格格式
ali = Alignment(horizontal='center', vertical='center',wrap_text=True )

#颜色填充
fill1 = PatternFill('solid', fgColor='E9E9F3')
fill2 = PatternFill('solid', fgColor='FFFFFF')

#边线和边框
sd1 = Side(style='thin', color='000000')
sd2 = Side(style='medium', color='000000')
border1 = Border(top=sd1, bottom=sd2, left=sd1, right=sd1)
border2 = Border(top=sd1, bottom=sd1, left=sd1, right=sd1)

#打包以上格式
sty1 = NamedStyle(name='sty1', font=font1, fill=fill1, border=border1, alignment=ali) #加粗、底色、粗底线
sty2 = NamedStyle(name='sty2', font=font2, fill=fill2, border=border2, alignment=ali) #其他文字
sty3 = NamedStyle(name='sty3', font=font1, fill=fill1, border=border2, alignment=ali) #加粗、底色
sty4 = NamedStyle(name='sty4', font=font1, fill=fill2, border=border2, alignment=ali) #加粗
sty5 = NamedStyle(name='sty5', font=font3, fill=fill2, alignment=ali) #黑体14号、加粗、无边框
#第一张表
rows1 = ws1.max_row
cols1 = ws1.max_column

for r in range(1, rows1+1):
    for c in range(1, cols1+1):
        if r == 1:
            ws1.cell(r, c).style = sty1
        else:
            ws1.cell(r, c).style = sty2
            
#ws1.column_dimensions['A'].width = 20.0  #调整列宽
#ws1.row_dimensions[1].height = 40  #调整行高
            
#第二张表
rows2 = ws2.max_row
cols2 = ws2.max_column 

ws2.insert_rows(1, 1)
ws2['A1'] = '本年度干部变动统计表'
ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols2)

rows2 = ws2.max_row
cols2 = ws2.max_column

for r in range(1, rows2+1):
    for c in range(1, cols2+1):
        if r == 1:
            ws2.cell(r, c).style = sty5
        elif ((r == 2)|(c == 1)):
            ws2.cell(r, c).style = sty3
        elif r == (rows2):
            ws2.cell(r, c).style = sty4
        else:
            ws2.cell(r, c).style = sty2
            
            
#第三张表
ws3.delete_rows(1, 1)
            
            
            
wb.save(r'C:\Users\User\Desktop\本年度干部变动统计-截止201911-2.xlsx')