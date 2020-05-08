# -*- coding: utf-8 -*-
"""
Created on Wed Nov 27 08:59:33 2019

@author: User
"""

input('即将对集团和证券公司干部信息明细表进行格式美化，按回车键继续... \n')

from openpyxl import load_workbook
from openpyxl.styles import Font, NamedStyle
from openpyxl.styles import PatternFill, Border, Side, Alignment

date = input('输入月度统计表的年月，(格式：YYYYMM):')
print('\n E:\\1-统计\\%s\\raw\\' %date)
input('请检查文件目录是否正确，确保目录下有以下文件：\n “干部信息明细表（数据清洗）.xlsx” \n 按回车键继续... \n')

wb = load_workbook(r'E:\1-统计\%s\raw\干部信息明细表（数据清洗）.xlsx' %date)

ws1 = wb['干部信息明细表总表']
ws2 = wb['集团和证券公司领导']
ws3 = wb['集团和证券公司总部干部']
ws4 = wb['分公司干部']
ws5 = wb['非一体化管控子公司']
ws6 = wb['营业部干部']
ws7 = wb['非行政职务']

#字体
font1 = Font(name='宋体', color='000000', size=10, b=True)
font2 = Font(name='宋体', color='000000', size=10)

#单元格格式
ali = Alignment(horizontal='center', vertical='center',wrap_text=True )

#颜色填充
fill1 = PatternFill('solid', fgColor='D9D9F3')
fill2 = PatternFill('solid', fgColor='FFFFFF')

#边线和边框
sd1 = Side(style='thin', color='000000')
sd2 = Side(style='medium', color='000000')
border1 = Border(top=sd1, bottom=sd2, left=sd1, right=sd1)
border2 = Border(top=sd1, bottom=sd1, left=sd1, right=sd1)

#打包以上格式
sty1 = NamedStyle(name='sty1', font=font1, fill=fill1,border=border1, alignment=ali)
sty2 = NamedStyle(name='sty2', font=font2, fill=fill2,border=border2, alignment=ali)

def setup(ws):
    ws.delete_cols(18, 14) #从18列开始删除，往后删14列（删除之前用pandas匹配的用于统计汇总的数据字段）

    rows = ws.max_row
    cols = ws.max_column

    #第一行表头和第二行开始的数据分别匹配格式
    for r in range(1, rows+1):
        for c in range(1, cols+1):
            if r == 1:
                ws.cell(r, c).style = sty1
            else:
                ws.cell(r, c).style = sty2

    #冻结首行
    ws.freeze_panes = 'A2'
    #设置列宽
    ws.column_dimensions["A"].width = 17
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 27
    ws.column_dimensions["H"].width = 11
    ws.column_dimensions["I"].width = 13
    ws.column_dimensions["J"].width = 9
    ws.column_dimensions["K"].width = 9
    ws.column_dimensions["L"].width = 10
    ws.column_dimensions["M"].width = 4
    ws.column_dimensions["O"].width = 4
    ws.column_dimensions["P"].width = 10
    ws.column_dimensions["Q"].width = 10

setup(ws1)
setup(ws2)
setup(ws3)
setup(ws4)
setup(ws5)
setup(ws6)
setup(ws7)

print('E:\\1-统计\\%s\\raw\\' %date)
input('将输出文件至上述目录，按回车键继续... \n')
wb.save(r'E:\1-统计\%s\raw\集团和证券公司干部信息明细表%s.xlsx' %(date,date))
